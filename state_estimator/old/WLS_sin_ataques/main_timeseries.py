from lib_timeseries import system_topology, system_measurements, system_constraints
import lib
import pandas as pd
from openpyxl import Workbook, load_workbook
import numpy as np

case = ['090neg', '090pos', '100pos']
workbook = Workbook()

excel_name = "Results_WLS_std2.xlsx"
file_name = 'std_2.json'
workbook.save(excel_name)




std_medidas = []
std_solucion = []
porcentaje_mejora = []

# Obtenemos la topología del parque
Nodes, Lines = system_topology('../data/pv_2_3.json')

# Para cada instante de tiempo
path = ['../data/pv_2_3_180_', '../data_Cati/pv_2_3_180_']
for c in case:
    for hour in ['08', '09', '10', '11', '12', '13', '14']:
        for minute in ['00', '15', '30', '45']:
            sheet_name = hour + '_' + minute + '_pf_' + c
            extended_path = [path[0] + hour + '_' + minute + '_pf_' + c + '/',
                             path[1] + hour + '_' + minute + '_pf_' + c + '/']
            print('Evaluando ' + hour + '_' + minute + '_pf_090neg')
            
            
            # Tomamos los valores de las medidas
            Meas, mjson, stdjson = system_measurements(extended_path, 
                                                       'measurements.json', 
                                                       file_name, 
                                                       Nodes, 
                                                       Lines, 
                                                       add_noise = True)
            # Construimos las restricciones
            Cons = system_constraints(Nodes)
    
            # Generamos la red y resolvemos el problema de estiamción de estado
            net = lib.grid(Nodes, Lines, Meas, Cons)
            res, sol, H, std_sol = net.state_estimation(tol = 1e-6, 
                                                        niter = 15, 
                                                        Huber = False, 
                                                        lmb = None, 
                                                        rn = True)
            net.norm_res()
            net.report()
            
            # Escribimos los resultados
            df = pd.DataFrame()
            df['Nombre'] = list(mjson.keys())
            df['Medidas'] = [mjson[item] if item[0] != 'I' else mjson[item]**2 for item in mjson]
            df['Std medidas'] = [item.std for item in net.meas]
            df['Medidas ruidosas'] = [item.value for item in net.meas]
            df['Estimación'] = [item.h() for item in net.meas]
            df['Residuos'] = [item.value - item.h() for item in net.meas]
            df['Std solucion'] = list(std_sol)
            df['Residuos normalizados'] = net.res_norm
            
            std_medidas.append(np.array([item.std for item in net.meas]))
            std_solucion.append(std_sol)
            porcentaje_mejora.append([(1 - (item[0] - item[1])/item[0])*100 for item in zip([item.std for item in net.meas], list(std_sol))])
            
            with pd.ExcelWriter(excel_name, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
# Resumen: std medidas min, max, avg // std solucion min. max, avg // % mejora min, max, avg
df = pd.DataFrame()
df['Std medidas min'] = np.min(np.array(std_medidas), axis = 0)
df['Std medidas max'] = np.max(np.array(std_medidas), axis = 0)
df['Std medidas avg'] = np.sum(np.array(std_medidas), axis = 0)/len(std_medidas)
df['Std solucion min'] = np.min(np.array(std_solucion), axis = 0)
df['Std solucion max'] = np.max(np.array(std_solucion), axis = 0)
df['Std solucion avg'] = np.sum(np.array(std_solucion), axis = 0)/len(std_solucion)
df['% mejora min'] = np.min(np.array(porcentaje_mejora), axis = 0)
df['% mejora max'] = np.max(np.array(porcentaje_mejora), axis = 0)
df['% mejora avg'] = np.sum(np.array(porcentaje_mejora), axis = 0)/len(porcentaje_mejora)

with pd.ExcelWriter(excel_name, engine="openpyxl", mode="a") as writer:
    df.to_excel(writer, sheet_name='Summary', index=False)    
    
book = load_workbook(excel_name)
summary_sheet = book['Summary']
sheets = [summary_sheet] + [book[sheet] for sheet in book.sheetnames if sheet != 'Summary']

    
new_book = Workbook()
for sheet in sheets:
    new_sheet = new_book.create_sheet(title=sheet.title)
    for row in sheet.iter_rows(values_only=True):
        new_sheet.append(row)

default_sheet = new_book['Sheet']
new_book.remove(default_sheet)

new_book.save(excel_name)
    
    
    
    
import json
data = {}
data['std_medidas'] = [item.tolist() for item in std_medidas]
data['std_solucion'] = [item.tolist() for item in std_solucion]
    
with open('data_' + file_name, 'w') as f:
    json.dump(data, f)
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    